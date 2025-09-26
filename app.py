import os
import re
import uuid
import threading
import json
import glob
from flask import Flask, render_template, request, send_from_directory, jsonify, flash, redirect, url_for, send_file
from werkzeug.utils import secure_filename
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string
from thefuzz import fuzz
from dadata import Dadata
import io
import value_dictionary_handler
import dictionary_matcher

# --- 1. Конфигурация приложения ---
TEMPLATES_DB_FOLDER = 'templates_db'
TEMPLATE_EXCEL_FOLDER = 'template_excel_files'
ALLOWED_EXTENSIONS = {'xlsx', 'xlsm'}

# --- НОВАЯ КОНФИГУРАЦИЯ DADATA ---
DADATA_API_KEY = "ВАШ_API_КЛЮЧ"
DADATA_SECRET_KEY = "ВАШ_СЕКРЕТНЫЙ_КЛЮЧ"

app = Flask(__name__)
app.config.from_mapping(
    TEMPLATES_DB_FOLDER=TEMPLATES_DB_FOLDER,
    TEMPLATE_EXCEL_FOLDER=TEMPLATE_EXCEL_FOLDER,
    SECRET_KEY='your-super-secret-key-change-it-for-production'
)

os.makedirs(TEMPLATES_DB_FOLDER, exist_ok=True)
os.makedirs(TEMPLATE_EXCEL_FOLDER, exist_ok=True)

# --- 2. Глобальное хранилище статусов задач ---
task_statuses = {}


# --- 3. Вспомогательные функции (без изменений) ---
def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


def normalize_header(header):
    if not isinstance(header, str):
        header = str(header)
    return re.sub(r'[\s\W_]+', '', header.lower())


def get_col_from_cell(cell_coord):
    if not cell_coord: return None
    match = re.match(r"([A-Z]+)", cell_coord.upper())
    return match.group(1) if match else None

@app.route('/value-dictionary')
def value_dictionary_ui():
    """Отображает страницу управления словарем значений."""
    rules = value_dictionary_handler.load_dictionary()
    return render_template('value_dictionary.html', rules=rules)

@app.route('/value-dictionary/add', methods=['POST'])
def add_to_value_dictionary():
    """Добавляет или обновляет правило в словаре значений."""
    canonical_word = request.form.get('canonical_word')
    find_words = request.form.get('find_words')
    if canonical_word and find_words:
        value_dictionary_handler.add_entry(canonical_word, find_words)
    return redirect(url_for('value_dictionary_ui'))

@app.route('/value-dictionary/delete', methods=['POST'])
def delete_from_value_dictionary():
    """Удаляет правило из словаря значений."""
    canonical_word = request.form.get('canonical_word')
    if canonical_word:
        value_dictionary_handler.delete_entry(canonical_word)
    return redirect(url_for('value_dictionary_ui'))


def _apply_value_dictionary(worksheet, task_id):
    """Применяет правила из словаря значений ко всем ячейкам листа."""
    # Получаем эффективную карту замен вида {'найти': 'заменить'}
    reverse_rules_map = value_dictionary_handler.get_reverse_lookup_map()
    if not reverse_rules_map:
        print("--> Словарь значений пуст, замена не требуется.")
        return

    task_statuses[task_id]['status'] = 'Выполняю замену по словарю значений...'
    print("--> Начата замена по словарю значений...")
    replacements_count = 0

    # Проходим по всем ячейкам на листе
    for row in worksheet.iter_rows():
        for cell in row:
            # Работаем только с ячейками, где есть строковые значения
            if cell.value and isinstance(cell.value, str):
                # Ищем точное совпадение в нашей карте
                replacement = reverse_rules_map.get(cell.value)
                if replacement is not None:
                    cell.value = replacement
                    replacements_count += 1

    print(f"--> Замена по словарю значений завершена. Сделано замен: {replacements_count}")
    task_statuses[task_id]['status'] = f'Замена по словарю завершена ({replacements_count} замен)'

def find_column_indices(worksheet, start_row, headers_to_find):
    indices = {}
    header_row = worksheet[start_row]
    normalized_map = {normalize_header(cell.value): cell.column for cell in header_row if cell.value}
    for key, target_header in headers_to_find.items():
        normalized_target = normalize_header(target_header)
        found_col = normalized_map.get(normalized_target)
        if found_col:
            indices[key] = found_col
    return indices


def apply_post_processing(task_id, workbook, start_row, function_name):
    if function_name == 'none' or not function_name:
        return

    worksheet = workbook.active
    dadata = Dadata(DADATA_API_KEY, DADATA_SECRET_KEY)
    total_rows = worksheet.max_row - start_row

    if function_name == 'coords_to_address':
        cols = find_column_indices(worksheet, start_row, {'lat': 'Широта', 'lon': 'Долгота', 'addr': 'Адрес'})
        if not all(k in cols for k in ['lat', 'lon', 'addr']):
            raise ValueError("Не найдены обязательные столбцы: 'Широта', 'Долгота', 'Адрес'")

        coords_to_process, rows_to_update = [], []
        for i, row_cells in enumerate(worksheet.iter_rows(min_row=start_row + 1)):
            lat_val, lon_val = row_cells[cols['lat'] - 1].value, row_cells[cols['lon'] - 1].value
            if lat_val and lon_val:
                coords_to_process.append({"lat": lat_val, "lon": lon_val})
                rows_to_update.append(start_row + 1 + i)

        if coords_to_process:
            try:
                task_statuses[task_id][
                    'status'] = f"Геокодирование (координаты->адрес): {len(coords_to_process)} адресов..."
                results = dadata.geolocate(name="address", queries=coords_to_process)
                for i, result in enumerate(results):
                    if result and result['suggestions']:
                        address = result['suggestions'][0]['value']
                        worksheet.cell(row=rows_to_update[i], column=cols['addr']).value = address
            except Exception as e:
                print(f"Ошибка геокодирования DaData: {e}")

    elif function_name == 'address_to_coords':
        cols = find_column_indices(worksheet, start_row, {'lat': 'Широта', 'lon': 'Долгота', 'addr': 'Адрес'})
        if not all(k in cols for k in ['lat', 'lon', 'addr']):
            raise ValueError("Не найдены обязательные столбцы: 'Широта', 'Долгота', 'Адрес'")

        addresses_to_process, rows_to_update = [], []
        for i, row_cells in enumerate(worksheet.iter_rows(min_row=start_row + 1)):
            addr_val = row_cells[cols['addr'] - 1].value
            if addr_val:
                addresses_to_process.append(str(addr_val))
                rows_to_update.append(start_row + 1 + i)

        if addresses_to_process:
            try:
                task_statuses[task_id][
                    'status'] = f"Геокодирование (адрес->координаты): {len(addresses_to_process)} адресов..."
                results = dadata.clean(name="address", source=addresses_to_process)
                for i, result in enumerate(results):
                    if result and result['geo_lat'] and result['geo_lon']:
                        worksheet.cell(row=rows_to_update[i], column=cols['lat']).value = float(result['geo_lat'])
                        worksheet.cell(row=rows_to_update[i], column=cols['lon']).value = float(result['geo_lon'])
            except Exception as e:
                print(f"Ошибка геокодирования DaData: {e}")


def _apply_manual_rules(source_ws, template_ws, rules, s_start_row, t_start_row, used_source_cols, used_template_cols):
    s_end_row = source_ws.max_row
    for rule in rules:
        s_col_letter = rule.get('s_col') or get_col_from_cell(rule.get('source_cell'))
        t_col_letter = rule.get('t_col') or rule.get('template_col')

        if not s_col_letter or not t_col_letter:
            continue

        s_col_idx, t_col_idx = column_index_from_string(s_col_letter), column_index_from_string(t_col_letter)

        if s_col_idx in used_source_cols or t_col_idx in used_template_cols:
            continue

        for i, row in enumerate(
                source_ws.iter_rows(min_row=s_start_row + 1, max_row=s_end_row, min_col=s_col_idx, max_col=s_col_idx)):
            source_cell = row[0]
            target_cell = template_ws.cell(row=t_start_row + 1 + i, column=t_col_idx)

            target_cell.value = source_cell.value
            if source_cell.hyperlink:
                target_cell.hyperlink = source_cell.hyperlink.target
                target_cell.style = "Hyperlink"

        used_source_cols.add(s_col_idx)
        used_template_cols.add(t_col_idx)


def _apply_dictionary_matching(source_ws, template_ws, s_start_row, t_start_row, used_source_cols, used_template_cols):
    reverse_dictionary = dictionary_matcher.get_reverse_dictionary()
    s_headers = {c.column: normalize_header(c.value) for c in source_ws[s_start_row] if c.value}
    t_headers = {c.column: normalize_header(c.value) for c in template_ws[t_start_row] if c.value}
    s_end_row = source_ws.max_row

    for s_col_idx, s_norm_h in s_headers.items():
        if s_col_idx in used_source_cols: continue
        canonical_name = reverse_dictionary.get(s_norm_h)
        if not canonical_name: continue
        normalized_canonical = normalize_header(canonical_name)

        for t_col_idx, t_norm_h in t_headers.items():
            if t_col_idx in used_template_cols: continue
            if t_norm_h == normalized_canonical:
                for i, row in enumerate(
                        source_ws.iter_rows(min_row=s_start_row + 1, max_row=s_end_row, min_col=s_col_idx,
                                            max_col=s_col_idx)):
                    source_cell = row[0]
                    target_cell = template_ws.cell(row=t_start_row + 1 + i, column=t_col_idx)

                    target_cell.value = source_cell.value
                    if source_cell.hyperlink:
                        target_cell.hyperlink = source_cell.hyperlink.target
                        target_cell.style = "Hyperlink"

                used_source_cols.add(s_col_idx)
                used_template_cols.add(t_col_idx)
                break


def _apply_auto_matching(source_ws, template_ws, s_start_row, t_start_row, used_source_cols, used_template_cols,
                         task_id):
    s_headers = {c.column: normalize_header(c.value) for c in source_ws[s_start_row] if c.value}
    t_headers = {c.column: normalize_header(c.value) for c in template_ws[t_start_row] if c.value}
    s_end_row = source_ws.max_row

    auto_source_headers = {k: v for k, v in s_headers.items() if k not in used_source_cols}
    auto_template_headers = {k: v for k, v in t_headers.items() if k not in used_template_cols}
    auto_matches = {}

    for s_col, s_norm in auto_source_headers.items():
        best_match, best_score = None, 0
        for t_col, t_norm in auto_template_headers.items():
            score = fuzz.ratio(s_norm, t_norm)
            if score > best_score:
                best_score, best_match = score, t_col
        if best_score > 75:
            auto_matches[s_col] = best_match
            auto_template_headers = {k: v for k, v in auto_template_headers.items() if k != best_match}

    rows_to_process = list(source_ws.iter_rows(min_row=s_start_row + 1, max_row=s_end_row))
    for i, row in enumerate(rows_to_process):
        for s_col, t_col in auto_matches.items():
            source_cell = row[s_col - 1]
            target_cell = template_ws.cell(row=t_start_row + 1 + i, column=t_col)

            target_cell.value = source_cell.value
            if source_cell.hyperlink:
                target_cell.hyperlink = source_cell.hyperlink.target
                target_cell.style = "Hyperlink"

        if (i + 1) % 100 == 0:
            task_statuses[task_id]['status'] = f'Автоматическое копирование: строка {i + 1}'


# --- 4. Основная функция обработки (ОБНОВЛЕНА) ---
def process_excel_hybrid(task_id, source_file_obj, template_file_obj, ranges, template_rules, private_rules,
                         post_function, original_template_filename):
    """
    Обрабатывает файлы Excel, используя потоки в памяти.
    """
    try:
        task_statuses[task_id] = {
            'progress': 5,
            'status': 'Подготовка...',
            'template_filename': original_template_filename
        }
        source_wb = load_workbook(filename=source_file_obj)
        source_ws = source_wb.active

        # <-- ИЗМЕНЕНО: Проверяем, является ли шаблон .xlsm, и используем флаг keep_vba=True -->
        is_macro_enabled = original_template_filename.lower().endswith('.xlsm')
        template_wb = load_workbook(filename=template_file_obj, keep_vba=is_macro_enabled)
        template_ws = template_wb.active

        s_start_row, t_start_row = ranges['s_start_row'], ranges['t_start_row']
        used_source_cols, used_template_cols = set(), set()

        task_statuses[task_id]['status'] = 'Выполняю частные правила...'
        _apply_manual_rules(source_ws, template_ws, private_rules, s_start_row, t_start_row, used_source_cols,
                            used_template_cols)
        task_statuses[task_id]['status'] = 'Применяю правила из шаблона...'
        _apply_manual_rules(source_ws, template_ws, template_rules, s_start_row, t_start_row, used_source_cols,
                            used_template_cols)
        task_statuses[task_id]['status'] = 'Проверяю по словарю...'
        _apply_dictionary_matching(source_ws, template_ws, s_start_row, t_start_row, used_source_cols,
                                   used_template_cols)
        task_statuses[task_id]['status'] = 'Ищу автоматические совпадения...'
        _apply_auto_matching(source_ws, template_ws, s_start_row, t_start_row, used_source_cols, used_template_cols,
                             task_id)
        _apply_value_dictionary(template_ws, task_id)
        task_statuses[task_id]['status'] = 'Запускаю пост-обработку...'
        apply_post_processing(task_id, template_wb, t_start_row, post_function)

        task_statuses[task_id]['status'] = 'Сохраняю результат...'
        processed_file_obj = io.BytesIO()
        template_wb.save(processed_file_obj)
        processed_file_obj.seek(0)
        source_wb.close()
        template_wb.close()
        task_statuses[task_id].update({'progress': 100, 'status': 'Готово!', 'result_file': processed_file_obj})

    except Exception as e:
        task_statuses[task_id].update({'progress': 100, 'status': f"Ошибка: {e}", 'result_file': None})


# --- Роуты Flask (без изменений) ---

@app.route('/templates')
def templates_list():
    template_files = glob.glob(os.path.join(app.config['TEMPLATES_DB_FOLDER'], '*.json'))
    templates_data = []
    for f in template_files:
        try:
            with open(f, 'r', encoding='utf-8') as file:
                data = json.load(file);
                data['id'] = os.path.basename(f).replace('.json', '');
                templates_data.append(data)
        except Exception as e:
            print(f"Ошибка чтения шаблона {f}: {e}")
    return render_template('templates_list.html', templates=templates_data)


@app.route('/templates/new')
def new_template_form():
    return render_template('create_template.html')


@app.route('/templates/create', methods=['POST'])
def create_template():
    try:
        template_name = request.form.get('template_name')
        header_start_cell = request.form.get('header_start_cell').upper()
        excel_file = request.files.get('excel_file')
        if not (template_name and header_start_cell and excel_file and excel_file.filename):
            flash("Ошибка: Название, начальная ячейка и Excel-файл шаблона должны быть заполнены.", "error");
            return redirect(url_for('new_template_form'))

        template_id = str(uuid.uuid4())
        _, file_extension = os.path.splitext(excel_file.filename)
        saved_excel_filename = f"{template_id}{file_extension}"
        excel_file.save(os.path.join(app.config['TEMPLATE_EXCEL_FOLDER'], saved_excel_filename))

        rules = []
        source_cells, template_cols = request.form.getlist('source_cell'), request.form.getlist('template_col')
        for i in range(len(source_cells)):
            if source_cells[i] and template_cols[i]: rules.append(
                {"source_cell": source_cells[i].upper(), "template_col": template_cols[i].upper()})

        template_data = {
            "template_name": template_name,
            "excel_file": saved_excel_filename,
            "original_filename": excel_file.filename,
            "header_start_cell": header_start_cell,
            "rules": rules
        }
        with open(os.path.join(app.config['TEMPLATES_DB_FOLDER'], f"{template_id}.json"), 'w', encoding='utf-8') as f:
            json.dump(template_data, f, ensure_ascii=False, indent=4)
        flash(f"Шаблон '{template_name}' успешно создан!", "success");
        return redirect(url_for('templates_list'))
    except Exception as e:
        flash(f"Произошла ошибка: {e}", "error");
        return redirect(url_for('new_template_form'))


@app.route('/templates/edit/<template_id>', methods=['GET', 'POST'])
def edit_template(template_id):
    print(f"--- ВХОД В ФУНКЦИЮ edit_template ---")
    print(f"Метод запроса: {request.method}")
    print(f"ID шаблона из URL: '{template_id}'")

    json_path = os.path.join(app.config['TEMPLATES_DB_FOLDER'], f"{template_id}.json")
    if not os.path.exists(json_path):
        print(f"!!! ШАБЛОН НЕ НАЙДЕН по пути: '{json_path}'")
        flash("Шаблон не найден.", "error")
        return redirect(url_for('templates_list'))

    if request.method == 'POST':
        print(f"--> Зашли в блок POST для ID: '{template_id}'")
        try:
            with open(json_path, 'r', encoding='utf-8') as f:
                template_data = json.load(f)

            # 1. Обновляем текстовые поля
            template_data['template_name'] = request.form.get('template_name')
            template_data['header_start_cell'] = request.form.get('header_start_cell').upper()

            # --- НОВЫЙ БЛОК: ОБРАБОТКА ЗАМЕНЫ ФАЙЛА ---
            new_excel_file = request.files.get('excel_file')
            if new_excel_file and new_excel_file.filename:
                print(f"--> Обнаружен новый файл для загрузки: '{new_excel_file.filename}'")
                if allowed_file(new_excel_file.filename):
                    # Обновляем имя файла в JSON.
                    # Примечание: сам файл физически не сохраняется,
                    # система запоминает только его имя, как и при создании.
                    template_data['excel_file'] = secure_filename(new_excel_file.filename)
                    print(f"--> Имя файла шаблона '{template_id}' будет заменено на '{template_data['excel_file']}'")
                else:
                    flash("Ошибка: Загруженный файл имеет недопустимый формат. Разрешены только .xlsx и .xlsm.", "error")
                    return redirect(url_for('edit_template', template_id=template_id))
            else:
                print("--> Новый файл не был загружен, имя файла останется прежним.")
            # --- КОНЕЦ НОВОГО БЛОКА ---

            # 3. Обновляем правила
            rules = []
            source_cells = request.form.getlist('source_cell')
            template_cols = request.form.getlist('template_col')
            for i in range(len(source_cells)):
                if source_cells[i] and template_cols[i]:
                    rules.append({
                        "source_cell": source_cells[i].upper(),
                        "template_col": template_cols[i].upper()
                    })
            template_data['rules'] = rules

            # 4. Сохраняем все изменения в JSON
            with open(json_path, 'w', encoding='utf-8') as f:
                json.dump(template_data, f, ensure_ascii=False, indent=4)

            print(f"--> POST-запрос успешен. Перенаправляем на templates_list.")
            flash(f"Шаблон '{template_data['template_name']}' успешно обновлен!", "success")
            return redirect(url_for('templates_list'))

        except Exception as e:
            print(f"!!! ОШИБКА в блоке POST: {e}")
            flash(f"Произошла ошибка при обновлении: {e}", "error")
            return redirect(url_for('edit_template', template_id=template_id))

    # Логика для GET-запроса (остается без изменений)
    print(f"--> Зашли в блок GET. Готовим страницу для ID: '{template_id}'")
    with open(json_path, 'r', encoding='utf-8') as f:
        template_data = json.load(f)

    return render_template('edit_template.html', template=template_data, template_id=template_id)


@app.route('/templates/download/<template_id>')
def download_template_file(template_id):
    json_path = os.path.join(app.config['TEMPLATES_DB_FOLDER'], f"{template_id}.json")
    if not os.path.exists(json_path):
        flash("Шаблон не найден.", "error")
        return redirect(url_for('templates_list'))

    with open(json_path, 'r', encoding='utf-8') as f:
        template_data = json.load(f)

    excel_filename = template_data.get('excel_file')
    original_filename = template_data.get('original_filename', 'template.xlsx')

    if not excel_filename or not os.path.exists(os.path.join(app.config['TEMPLATE_EXCEL_FOLDER'], excel_filename)):
        flash("Для этого шаблона не найден файл Excel.", "error")
        return redirect(url_for('edit_template', template_id=template_id))

    return send_from_directory(
        app.config['TEMPLATE_EXCEL_FOLDER'],
        excel_filename,
        as_attachment=True,
        download_name=original_filename
    )


@app.route('/dictionary')
def dictionary_ui():
    return render_template('dictionary.html', dictionary=dictionary_matcher.load_dictionary())


@app.route('/dictionary/add', methods=['POST'])
def add_to_dictionary():
    # --- НАЧАЛО БЛОКА ЛОГИРОВАНИЯ ---
    print("\n--- [СЛОВАРЬ] ПОЛУЧЕН POST-ЗАПРОС НА /dictionary/add ---")
    try:
        # Получаем данные из формы
        canonical_name = request.form.get('canonical_name')
        synonyms = request.form.get('synonyms', '')

        print(f"  [ЛОГ] Получено основное имя: '{canonical_name}' (Тип: {type(canonical_name)})")
        print(f"  [ЛОГ] Получены синонимы: '{synonyms}' (Тип: {type(synonyms)})")

        # Проверяем, есть ли основное имя
        if canonical_name:
            print("  [ЛОГ] Основное имя присутствует. Вызываю dictionary_matcher.add_entry...")
            # Вызываем функцию для добавления/обновления записи
            dictionary_matcher.add_entry(canonical_name, synonyms)
            print("  [ЛОГ] Вызов dictionary_matcher.add_entry ЗАВЕРШЕН.")
            flash(f"Запись '{canonical_name}' успешно сохранена.", "success")
        else:
            # Если по какой-то причине основное имя не пришло
            print("  [ЛОГ] ОШИБКА: Основное имя (canonical_name) не было получено из формы.")
            flash("Ошибка: Не удалось сохранить запись, так как основное имя не было передано.", "error")

    except Exception as e:
        # Логируем любую непредвиденную ошибку
        print(f"  [ЛОГ] КРИТИЧЕСКАЯ ОШИБКА в /dictionary/add: {e}")
        flash(f"Произошла критическая ошибка: {e}", "error")

    print("--- [СЛОВАРЬ] ЗАВЕРШЕНИЕ ЗАПРОСА. Перенаправление на /dictionary ---\n")
    # --- КОНЕЦ БЛОКА ЛОГИРОВАНИЯ ---

    return redirect(url_for('dictionary_ui'))


@app.route('/dictionary/delete', methods=['POST'])
def delete_from_dictionary():
    canonical_name = request.form.get('canonical_name')
    if canonical_name: dictionary_matcher.delete_entry(canonical_name)
    return redirect(url_for('dictionary_ui'))


@app.route('/')
def index():
    template_files = glob.glob(os.path.join(app.config['TEMPLATES_DB_FOLDER'], '*.json'))
    templates_data = []
    for f in template_files:
        try:
            with open(f, 'r', encoding='utf-8') as file:
                data = json.load(file);
                data['id'] = os.path.basename(f).replace('.json', '');
                templates_data.append(data)
        except Exception as e:
            print(f"Ошибка чтения шаблона {f}: {e}")
    return render_template('index.html', saved_templates=templates_data)


@app.route('/process', methods=['POST'])
def process_files():
    try:
        source_file = request.files.get('source_file')
        if not (source_file and source_file.filename and allowed_file(source_file.filename)):
            return jsonify({'error': 'Исходный файл .xlsx или .xlsm должен быть загружен.'}), 400

        source_file_obj = io.BytesIO(source_file.read())

        original_template_filename = ''
        template_rules, selected_template_id = [], request.form.get('saved_template')

        if selected_template_id:
            json_path = os.path.join(app.config['TEMPLATES_DB_FOLDER'], f"{selected_template_id}.json")
            with open(json_path, 'r', encoding='utf-8') as f:
                template_info = json.load(f)

            excel_filename = template_info.get('excel_file')
            original_template_filename = template_info.get('original_filename', 'template.xlsx')
            template_file_path = os.path.join(app.config['TEMPLATE_EXCEL_FOLDER'], excel_filename)

            if excel_filename and os.path.exists(template_file_path):
                 with open(template_file_path, 'rb') as f:
                    template_file_obj = io.BytesIO(f.read())
            else:
                from openpyxl import Workbook
                template_wb = Workbook()
                template_file_obj = io.BytesIO()
                template_wb.save(template_file_obj)
                template_file_obj.seek(0)

            t_start_cell = template_info['header_start_cell']
            template_rules = template_info.get('rules', [])
        else:
            template_file = request.files.get('template_file')
            if not (template_file and template_file.filename and allowed_file(template_file.filename)):
                return jsonify({'error': 'Если шаблон не выбран, его нужно загрузить вручную (.xlsx или .xlsm).'}), 400

            original_template_filename = template_file.filename
            template_file_obj = io.BytesIO(template_file.read())
            t_start_cell = request.form.get('template_range_start').upper()

        ranges = {
            's_start_row': int(re.search(r'\d+', request.form.get('source_range_start').upper()).group()),
            't_start_row': int(re.search(r'\d+', t_start_cell).group())
        }
        private_rules = []
        s_cols, t_cols = request.form.getlist('manual_source_col'), request.form.getlist('manual_template_col')
        for i in range(len(s_cols)):
            if s_cols[i] and t_cols[i]: private_rules.append({'s_col': s_cols[i].upper(), 't_col': t_cols[i].upper()})

        post_function = request.form.get('post_processing_function', 'none')
        task_id = str(uuid.uuid4())

        thread = threading.Thread(target=process_excel_hybrid, args=(
            task_id, source_file_obj, template_file_obj, ranges, template_rules, private_rules, post_function, original_template_filename))
        thread.start()
        return jsonify({'task_id': task_id})
    except Exception as e:
        return jsonify({'error': f'Ошибка на сервере: {e}'}), 500


@app.route('/status/<task_id>')
def task_status(task_id):
    status_info = task_statuses.get(task_id, {})
    if status_info.get('result_file'):
        template_filename = status_info.get('template_filename', 'template.xlsx')
        _, file_extension = os.path.splitext(template_filename)
        file_name = f"processed_{task_id}{file_extension}"

        return jsonify({'status': status_info['status'], 'progress': status_info['progress'], 'result_file': file_name})
    return jsonify(status_info)


@app.route('/download/<filename>')
def download_file(filename):
    task_id_with_ext = filename.replace('processed_', '')
    task_id, file_extension = os.path.splitext(task_id_with_ext)

    status_info = task_statuses.get(task_id)
    if status_info and status_info.get('result_file'):
        file_obj = status_info['result_file']

        mimetype = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        if file_extension.lower() == '.xlsm':
            mimetype = 'application/vnd.ms-excel.sheet.macroEnabled.12'

        return send_file(file_obj, as_attachment=True, download_name=filename,
                         mimetype=mimetype)
    return "Файл не найден или обработка еще не завершена.", 404


def get_cell_content(cell):
    if cell.hyperlink and cell.hyperlink.target:
        return f"{cell.value} ({cell.hyperlink.target})"
    return cell.value


@app.route('/templates/delete/<template_id>', methods=['POST'])
def delete_template(template_id):
    """Удаляет JSON-файл шаблона по его ID."""
    try:
        # Формируем полный путь к файлу шаблона
        json_path = os.path.join(app.config['TEMPLATES_DB_FOLDER'], f"{secure_filename(template_id)}.json")

        print(f"--> Поступил запрос на удаление шаблона: '{template_id}'")
        print(f"--> Путь к файлу для удаления: '{json_path}'")

        if os.path.exists(json_path):
            # Если файл существует, удаляем его
            os.remove(json_path)
            flash(f"Шаблон успешно удален.", "success")
            print(f"--> Файл '{json_path}' успешно удален.")
        else:
            # Если файл по какой-то причине не найден
            flash("Ошибка: Шаблон для удаления не найден.", "error")
            print(f"!!! ОШИБКА: Файл '{json_path}' для удаления не найден.")

    except Exception as e:
        # Обработка любых других ошибок
        flash(f"Произошла ошибка при удалении шаблона: {e}", "error")
        print(f"!!! КРИТИЧЕСКАЯ ОШИБКА при удалении шаблона '{template_id}': {e}")

    # Возвращаем пользователя на страницу со списком шаблонов
    return redirect(url_for('templates_list'))


if __name__ == '__main__':
    app.run(debug=True, port=5012)